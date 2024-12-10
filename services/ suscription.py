import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import Calendar
from datetime import datetime
from dateutil.relativedelta import relativedelta

# Global variables
file_name = "gym_management.xlsx"
df_subscription = pd.DataFrame(columns=["ID", "Subscription Date", "Duration", "Ending Date"])
df_payment = pd.DataFrame(columns=["ID", "Amount", "Advance Pay", "Remaining"])
df_lockers = pd.DataFrame(columns=["Locker Number", "Available", "ID"])

# Load or create workbook and data
def load_data_from_excel():
    global df_subscription, df_payment, df_lockers
    
    try:
        wb = load_workbook(file_name)
    except FileNotFoundError:
        wb = Workbook()
        wb.save(file_name)  # Create a new file if it doesn't exist

    # Load subscription data
    if "Subscription" in wb.sheetnames:
        ws = wb["Subscription"]
        data = list(ws.values)
        df_subscription = pd.DataFrame(data[1:], columns=data[0])  # Skip first row and use it as header

    # Load payment data
    if "Payment" in wb.sheetnames:
        ws = wb["Payment"]
        data = list(ws.values)
        df_payment = pd.DataFrame(data[1:], columns=data[0])  # Skip first row and use it as header

    # Load locker data
    if "Lockers" in wb.sheetnames:
        ws = wb["Lockers"]
        data = list(ws.values)
        df_lockers = pd.DataFrame(data[1:], columns=data[0])  # Skip first row and use it as header
        df_lockers["Locker Number"] = pd.to_numeric(df_lockers["Locker Number"], errors="coerce").astype("Int64")
        df_lockers["Available"] = df_lockers["Available"].astype(bool)
        df_lockers["ID"] = pd.to_numeric(df_lockers["ID"], errors="coerce").astype("Int64")
    else:
        # If the lockers sheet doesn't exist, initialize it with proper values
        df_lockers = pd.DataFrame({
            "Locker Number": range(1, 11),
            "Available": [True] * 10,
            "ID": [None] * 10
        })

    return wb

# Save data to Excel
def save_to_excel(wb):
    sheets = {
        "Subscription": df_subscription,
        "Payment": df_payment,
        "Lockers": df_lockers
    }

    for sheet_name, df in sheets.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.delete_rows(1, ws.max_row)  # Clear previous data, but keep the headers
        else:
            ws = wb.create_sheet(title=sheet_name)

        # Write headers and data
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

    wb.save(file_name)

# Initialize lockers
def initialize_lockers():
    global df_lockers

    # Ensure lockers data is loaded from the sheet
    if df_lockers.empty:
        df_lockers = pd.DataFrame({
            "Locker Number": range(1, 11),
            "Available": [True] * 10,
            "ID": [None] * 10
        })

# Assign locker
def assign_locker(subscriber_id):
    global df_lockers

    available_lockers = df_lockers[df_lockers["Available"] == True]
    if available_lockers.empty:
        messagebox.showerror("Error", "No lockers available!")
        return None

    locker_index = available_lockers.index[0]
    locker_number = df_lockers.loc[locker_index, "Locker Number"]
    df_lockers.loc[locker_index, "Available"] = False
    df_lockers.loc[locker_index, "ID"] = subscriber_id
    return locker_number

# Update locker manually
def update_locker_status(locker_number, subscriber_id):
    global df_lockers

    if locker_number not in df_lockers["Locker Number"].values:
        messagebox.showerror("Error", f"Locker {locker_number} does not exist!")
        return

    locker_index = df_lockers.index[df_lockers["Locker Number"] == locker_number].tolist()[0]
    if df_lockers.loc[locker_index, "Available"]:
        df_lockers.loc[locker_index, "Available"] = False
        df_lockers.loc[locker_index, "ID"] = subscriber_id
        messagebox.showinfo("Success", f"Locker {locker_number} assigned to ID {subscriber_id}.")
    else:
        messagebox.showwarning("Unavailable", f"Locker {locker_number} is already occupied!")

# Change locker GUI
def change_locker_status_gui():
    def submit_locker_change():
        try:
            locker_number = int(locker_number_entry.get())
            subscriber_id = int(subscriber_id_entry.get())
            update_locker_status(locker_number, subscriber_id)
            locker_window.destroy()
        except ValueError:
            messagebox.showerror("Invalid Input", "Please enter valid numbers for locker and ID!")

    locker_window = tk.Toplevel(root)
    locker_window.title("Change Locker Status")
    locker_window.geometry("400x200")

    ttk.Label(locker_window, text="Locker Number:").grid(row=0, column=0, padx=10, pady=10)
    locker_number_entry = ttk.Entry(locker_window, width=20)
    locker_number_entry.grid(row=0, column=1, padx=10, pady=10)

    ttk.Label(locker_window, text="Subscriber ID:").grid(row=1, column=0, padx=10, pady=10)
    subscriber_id_entry = ttk.Entry(locker_window, width=20)
    subscriber_id_entry.grid(row=1, column=1, padx=10, pady=10)

    ttk.Button(locker_window, text="Update Locker", command=submit_locker_change).grid(row=2, column=0, columnspan=2, pady=20)

# Calculate end date
def calculate_end_date(start_date, duration):
    try:
        start_date = datetime.strptime(start_date, "%Y-%m-%d")
        end_date = start_date + relativedelta(months=duration)
        return end_date.strftime("%Y-%m-%d")
    except ValueError:
        messagebox.showerror("Invalid Input", "Please select a valid start date.")
        return None

# Add subscription
def add_subscription():
    try:
        subscription_id = len(df_subscription) + 1
        subscription_date = start_date_entry.get()
        duration = selected_duration.get()
        end_date = calculate_end_date(subscription_date, duration)
        amount = int(amount_entry.get())
        advance = int(advance_entry.get())
        remaining = amount - advance

        df_subscription.loc[len(df_subscription)] = [subscription_id, subscription_date, duration, end_date]
        df_payment.loc[len(df_payment)] = [subscription_id, amount, advance, remaining]

        locker_number = assign_locker(subscription_id)
        if locker_number is None:
            return

        save_to_excel(wb)
        messagebox.showinfo("Success", f"Subscription added with ID {subscription_id} and Locker {locker_number}.")
    except Exception as e:
        messagebox.showerror("Error", str(e))

# Get date with calendar
def get_date():
    def on_select():
        selected_date = cal.get_date()
        start_date_entry.delete(0, tk.END)
        start_date_entry.insert(0, selected_date)
        date_window.destroy()

    date_window = tk.Toplevel(root)
    date_window.title("Select a Date")
    cal = Calendar(date_window, selectmode="day", date_pattern="yyyy-mm-dd")
    cal.pack(pady=20)
    tk.Button(date_window, text="Select Date", command=on_select).pack()

# Load existing data from Excel
wb = load_data_from_excel()

# Tkinter GUI
root = tk.Tk()
root.geometry("600x400")
root.title("Gym Management System")

frame = tk.Frame(root)
frame.pack(pady=10)

tk.Label(frame, text="Starting Date").grid(row=0, column=0)
start_date_entry = ttk.Entry(frame)
start_date_entry.grid(row=0, column=1)
ttk.Button(frame, text="Select Date", command=get_date).grid(row=0, column=2)

tk.Label(frame, text="Duration (Months)").grid(row=1, column=0)
selected_duration = tk.IntVar(value=1)
for i, duration in enumerate([1, 3, 6, 12]):
    ttk.Radiobutton(frame, text=str(duration), value=duration, variable=selected_duration).grid(row=1, column=i+1)

tk.Label(frame, text="Amount").grid(row=2, column=0)
amount_entry = ttk.Entry(frame)
amount_entry.grid(row=2, column=1)

tk.Label(frame, text="Advance Pay").grid(row=3, column=0)
advance_entry = ttk.Entry(frame)
advance_entry.grid(row=3, column=1)

ttk.Button(frame, text="Add Subscription", command=add_subscription).grid(row=4, column=0, columnspan=3, pady=10)
ttk.Button(frame, text="Change Locker Status", command=change_locker_status_gui).grid(row=5, column=0, columnspan=2, pady=10)

root.mainloop()
